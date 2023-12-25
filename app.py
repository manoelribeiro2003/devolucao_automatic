from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait #para trabalhar com elementos que talvez ainda nao existam no DOM
from selenium.webdriver.support import expected_conditions as EC #esperar pela presença de um elemento
import time
from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

#cria/carrega webdriver
pagina = 'https://amberlive.000webhostapp.com/pagina.html'
driver = webdriver.Firefox()
driver.get(pagina)
    
#carregar o workbook e a worksheet
arquivo = 'devolucao.xlsx'
wb = load_workbook(arquivo)
ws = wb.active

        
def get_codigos_produtos():
    array_cod_prod = []
    #guardar o array de codigos de produtos
    for row in range(2, 6):
        array_cod_prod.append(ws['A'+str(row)].value)
    return array_cod_prod

def get_nomes_produtos():
    array_nome_prod = []
    #guardar o array de nomes de produtos
    for row in range(2, 6):
        array_nome_prod.append(ws['B'+str(row)].value)
    return array_nome_prod
        
def get_quantidades():
    array_quant_result = []
    #guardar o array de quantidades resultantes de produtos
    for row in range(2, 6):
        array_quant_result.append(int(ws['C'+str(row)].value)-int(ws['F'+str(row)].value))
    return array_quant_result

def get_lotes_produtos():
    array_lote = []
    #guardar o array de lotes de produtos
    for row in range(2, 6):
        array_lote.append(ws['D'+str(row)].value)
    return array_lote

def inserir():
    #clico no campo Produto
    prod = driver.find_element(By.ID, "Product_Name")
    prod.click()
    time.sleep(0.2)
    #digito o codigo e aperto Enter
    prod.send_keys([linha]+Keys.RETURN)
    #seleciono o botao do endereço de origem
    xpath_orig = '//button[@class="pui-button ui-widget ui-state-default ui-corner-right pui-button-icon-only"]'
    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, xpath_orig))
    )
    orig = driver.find_elements(By.XPATH, xpath_orig)
    orig[2].click()
    #escolho dentre os elementos (divs) o produto com base no lote
    xpath_lote = str("//div[@style=\"padding-left: 10px;margin-top: -10px;\" and contains(., 'Lote: {}')].".format(get_lotes_produtos()[linha]))
    lote = driver.find_element(By.XPATH, xpath_lote)
    lote.click()
    #escrevo a quantidade certa
    id_quant = driver.find_element(By.ID, "Ammount")
    id_quant.clear()
    id_quant.send_keys(array_quant_result[linha])
    #clico na posicao da devolucao
    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, "elemento da posicao"))
    )
    elemento_posic = driver.find_element(By.XPATH, "elemento da posicao")
    elemento_posic.click()
    
def confirmar():
    #e clico no botao verde
    botao_verde = driver.find_element(By.ID, "MovingProduct_Submit")
    botao_verde.click()
    
def prox():
#controle de iteração
    while var_control:
        var_control = input("Digite o valor: ")
        print(type(var_control))






#guardar o array de codigos de produtos
for row in range(2, 6):
    array_cod_prod.append(ws['A'+str(row)].value)

#guardar o array de nomes de produtos
for row in range(2, 6):
    array_nome_prod.append(ws['B'+str(row)].value)

#guardar o array de quantidades resultantes de produtos
for row in range(2, 6):
    array_quant_result.append(int(ws['C'+str(row)].value)-int(ws['F'+str(row)].value))

#guardar o array de lotes de produtos
for row in range(2, 6):
    array_lote.append(ws['D'+str(row)].value)


tamanho_devolucao = 0
linha = 2

#saber o tamanho da devolucao
while ws['A'+str(linha)].value != None:
    linha +=1
    tamanho_devolucao +=1
    print(tamanho_devolucao)
    if ws['A'+str(linha)].value == None:
        break
print("O tamanho da devolucao é: ",tamanho_devolucao)
    
#clico na opcao de devolucao
WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.ID, "OriginalAddress_Devolution"))
    )
elem = driver.find_element(By.ID, "OriginalAddress_Devolution")
elem.click()
time.sleep(0.3)    

#logica da devolucao
for linha in tamanho_devolucao:
    var_control = True
    
    #clico no campo Produto
    prod = driver.find_element(By.ID, "Product_Name")
    prod.click()
    time.sleep(0.2) 

    #digito o codigo e aperto Enter
    prod.send_keys(array_cod_prod[linha]+Keys.RETURN)

    #seleciono o botao do endereço de origem
    xpath_orig = '//button[@class="pui-button ui-widget ui-state-default ui-corner-right pui-button-icon-only"]'
    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, xpath_orig))
    )
    orig = driver.find_elements(By.XPATH, xpath_orig)
    orig[2].click()

    #escolho dentre os elementos (divs) o produto com base no lote
    xpath_lote = str("//div[@style=\"padding-left: 10px;margin-top: -10px;\" and contains(., 'Lote: {}')].".format(array_lote[linha]))
    lote = driver.find_element(By.XPATH, xpath_lote)
    lote.click()

    #escrevo a quantidade certa
    id_quant = driver.find_element(By.ID, "Ammount")
    id_quant.clear()
    id_quant.send_keys(array_quant_result[linha])

    #clico na posicao da devolucao
    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, "elemento da posicao"))
    )
    elemento_posic = driver.find_element(By.XPATH, "elemento da posicao")
    elemento_posic.click()


    #e clico no botao verde
    botao_verde = driver.find_element(By.ID, "MovingProduct_Submit")
    botao_verde.click()
    
    #controle de iteração
    while var_control:
        var_control = input("Digite o valor: ")
        print(type(var_control))
    

    
    
# # #clico na opcao de devolucao
# elem = driver.find_element(By.ID, "OriginalAddress_Devolution")
# elem.click()

# time.sleep(1)

# # #clico no campo Produto
# prod = driver.find_element(By.ID, "Product_Name")
# prod.click()

# # #digito o codigo e aperto Enter
# prod.send_keys(array_cod_prod[0]+Keys.RETURN)

# # #seleciono o botao do endereço de origem
# xpath_orig = '//button[@class="pui-button ui-widget ui-state-default ui-corner-right pui-button-icon-only"]'
# WebDriverWait(driver, 5).until(
#     EC.presence_of_element_located((By.XPATH, xpath_orig))
# )
# orig = driver.find_elements(By.XPATH, xpath_orig)
# orig[2].click()

# # #escolho dentre os elementos (divs) o produto com base no lote
# xpath_lote = "//div[@style=\"padding-left: 10px;margin-top: -10px;\" and contains(., 'Lote: ')]"
# lote = driver.find_element(By.XPATH, xpath_lote + array_lote[0])
# lote.click()

# # #escrevo a quantidade certa
# id_quant = driver.find_element(By.ID, "Ammount")
# id_quant.clear()
# id_quant.send_keys(array_quant_result[0])

# # #seleciono o botao para mostrar a posicao de devolucao
# # # -> no desenvolvimento do codigo, haviam dois botões 
# # # (talvez devido a não renderização, olhar foto de registro name: "foto_botao_posicao.png". 
# # # Deve ser devido ao estado do botão - ativo/não ativo)

# # #clico na posicao da devolucao
# WebDriverWait(driver, 5).until(
#     EC.presence_of_element_located((By.XPATH, "elemento da posicao"))
# )
# elemento_posic = driver.find_element(By.XPATH, "elemento da posicao")
# elemento_posic.click()


# # #e clico no botao verde
# botao_verde = driver.find_element(By.ID, "MovingProduct_Submit")
# botao_verde.click()